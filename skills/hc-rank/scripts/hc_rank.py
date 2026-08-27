#!/usr/bin/env python3
"""
HC Rank — Common Crawl Harmonic Centrality lookup (bulk-capable, stdlib-first).

Reads Common Crawl's published domain-level Web Graph ranks file directly, so it
never depends on the (currently erroring) webgraph.metehan.ai front-end and scales
to bulk lists. Harmonic Centrality measures a domain's *proximity to the web's link
core*; Common Crawl uses it to set crawl priority, which drives how much of a site
enters LLM training data.

Workflow:
    python hc_rank.py build                       # one-time per release (~2.35GB download)
    python hc_rank.py lookup google.com bbc.co.uk
    python hc_rank.py bulk prospects.txt --xlsx prospects_hc.xlsx
    python hc_rank.py releases                     # list available web-graph releases

Design: pure-stdlib by default (gzip + csv) so it runs with ZERO external installs,
including on bleeding-edge Pythons without duckdb wheels. If `duckdb` is importable,
`build` also writes a Parquet for sub-second repeat lookups; otherwise queries stream
the cached .gz (one pass per run — fine for audit / link-building batches).

Data: https://data.commoncrawl.org/projects/hyperlinkgraph/<release>/domain/
      <release>-domain-ranks.txt.gz
Columns (TSV, sorted by harmonicc_pos asc):
      harmonicc_pos  harmonicc_val  pr_pos  pr_val  host_rev  n_hosts
      host_rev is the REVERSED registered domain, e.g. com.google, uk.co.bbc
"""

import argparse
import csv
import gzip
import json
import os
import sys
import time
from pathlib import Path

# Latest known domain-level web graph release (verified 2026-06-02).
# Run `hc_rank.py releases` to discover newer ones, then `build --release <id>`.
LATEST_RELEASE = "cc-main-2026-mar-apr-may"
BASE = "https://data.commoncrawl.org/projects/hyperlinkgraph"
WEBGRAPH_INDEX = "https://index.commoncrawl.org/web-graphs-index.html"

# Grade bands by HC rank position (1 = most central). Documented heuristic.
GRADE_BANDS = [
    (10_000,       "A+ Web core"),
    (100_000,      "A  Near-core"),
    (1_000_000,    "B  Strong"),
    (10_000_000,   "C  Mid"),
    (50_000_000,   "D  Peripheral"),
    (float("inf"), "E  Edge"),
]

# Fallback multi-label public suffixes (used only when `tldextract` isn't installed).
TWO_LABEL_SUFFIXES = {
    "co.uk", "org.uk", "gov.uk", "ac.uk", "me.uk", "ltd.uk", "plc.uk", "net.uk",
    "com.au", "net.au", "org.au", "edu.au", "gov.au", "co.nz", "org.nz", "co.za",
    "co.jp", "or.jp", "ne.jp", "com.br", "net.br", "com.mx", "co.in", "net.in",
    "org.in", "com.sg", "com.hk", "co.kr", "com.tr", "com.cn", "co.il", "com.ar",
    "com.tw", "com.ua", "co.id", "com.my", "com.ph", "com.pk", "com.sa", "co.th",
}


# --------------------------------------------------------------------------- cache
def cache_dir() -> Path:
    d = Path(os.environ.get("CC_WEBGRAPH_DIR", Path.home() / ".cache" / "cc-webgraph"))
    d.mkdir(parents=True, exist_ok=True)
    return d


def meta_path() -> Path:
    return cache_dir() / "meta.json"


def load_meta() -> dict:
    p = meta_path()
    if not p.exists():
        sys.exit("No index built yet. Run:  python hc_rank.py build")
    return json.loads(p.read_text())


# ------------------------------------------------------------------- domain keys
def registered_domain(domain: str) -> str:
    """bbc.co.uk -> bbc.co.uk ; https://www.bbc.co.uk/news -> bbc.co.uk ; blog.x.com -> x.com"""
    host = domain.strip().lower()
    if "//" in host:
        host = host.split("//", 1)[1]
    host = host.split("/")[0].split("?")[0].split("#")[0].split(":")[0]
    host = host.strip(".")
    if not host:
        return ""
    try:  # prefer full Public Suffix List correctness when available
        import tldextract
        ext = tldextract.extract(host)
        # `registered_domain` was renamed to `top_domain_under_public_suffix`; support both.
        reg = getattr(ext, "top_domain_under_public_suffix", None) or ext.registered_domain
        if reg:
            return reg
    except Exception:
        pass
    labels = host.split(".")
    if len(labels) <= 2:
        return host
    last2 = ".".join(labels[-2:])
    if last2 in TWO_LABEL_SUFFIXES and len(labels) >= 3:
        return ".".join(labels[-3:])
    return last2


def reverse_domain(domain: str) -> str:
    reg = registered_domain(domain)
    return ".".join(reversed(reg.split("."))) if reg else ""


def grade(pos: int) -> str:
    for ceiling, label in GRADE_BANDS:
        if pos <= ceiling:
            return label
    return "E  Edge"


# ------------------------------------------------------------------------- build
def ranks_url(release: str) -> str:
    return f"{BASE}/{release}/domain/{release}-domain-ranks.txt.gz"


def download_resumable(url: str, dest: Path):
    import urllib.request
    head = urllib.request.urlopen(urllib.request.Request(url, method="HEAD"), timeout=60)
    total = int(head.headers.get("Content-Length", 0))
    pos = dest.stat().st_size if dest.exists() else 0
    if total and pos == total:
        print(f"Already downloaded: {dest.name} ({total/1e9:.2f} GB)", file=sys.stderr)
        return
    if pos > total:
        pos = 0
    req = urllib.request.Request(url, headers={"Range": f"bytes={pos}-"} if pos else {})
    mode = "ab" if pos else "wb"
    print(f"Downloading {url}\n  -> {dest}  ({total/1e9:.2f} GB){' [resume]' if pos else ''}",
          file=sys.stderr)
    t0 = time.time()
    with urllib.request.urlopen(req, timeout=120) as r, open(dest, mode) as f:
        while True:
            chunk = r.read(1 << 20)
            if not chunk:
                break
            f.write(chunk)
            pos += len(chunk)
            if total:
                pct = 100 * pos / total
                mbps = (pos / 1e6) / max(time.time() - t0, 1e-6)
                print(f"\r  {pct:5.1f}%  {pos/1e9:5.2f}/{total/1e9:.2f} GB  {mbps:6.1f} MB/s",
                      end="", file=sys.stderr)
    print(file=sys.stderr)


def iter_rows(gz_path: Path):
    """Yield (harmonicc_pos, harmonicc_val, pr_pos, pr_val, host_rev, n_hosts), skipping header."""
    with gzip.open(gz_path, "rt", encoding="utf-8", errors="replace") as fh:
        rd = csv.reader(fh, delimiter="\t")
        for row in rd:
            if not row or row[0].startswith("#"):
                continue
            if len(row) < 6:
                continue
            yield row


def _duckdb_usable() -> bool:
    """duckdb is opt-in and version-gated: it SEGFAULTS on CPython 3.14 (uncatchable),
    which previously killed `build` before meta.json was written. Skip it there and
    whenever HC_NO_DUCKDB is set. Streaming mode is always correct, just slower."""
    if os.environ.get("HC_NO_DUCKDB", "").lower() in ("1", "true", "yes"):
        return False
    if sys.version_info[:2] >= (3, 14):
        return False
    try:
        import duckdb  # noqa: F401
    except Exception:
        return False
    return True


def _parquet_copy(gz: Path, parquet: Path):
    """The actual duckdb COPY. Runs in a SUBPROCESS (see try_build_parquet) so a
    native crash cannot take down the parent build."""
    import duckdb
    con = duckdb.connect()
    con.execute(
        """
        COPY (SELECT * FROM read_csv(?, delim='\t', header=false, skip=1, compression='gzip',
            columns={'harmonicc_pos':'BIGINT','harmonicc_val':'DOUBLE','pr_pos':'BIGINT',
                     'pr_val':'DOUBLE','host_rev':'VARCHAR','n_hosts':'BIGINT'}))
        TO ? (FORMAT parquet, COMPRESSION zstd)
        """,
        [str(gz), str(parquet)],
    )
    con.close()


def try_build_parquet(gz: Path, parquet: Path) -> bool:
    """Best-effort Parquet build for fast repeat lookups. Crash-isolated: the duckdb
    COPY runs in a child process, so a segfault (return code != 0 / killed) just leaves
    us in streaming mode instead of stranding the whole index."""
    if not _duckdb_usable():
        print("duckdb skipped (unusable on this Python / HC_NO_DUCKDB) — streaming mode.",
              file=sys.stderr)
        return False
    import subprocess
    print(f"duckdb present — building Parquet in isolated subprocess: {parquet.name}",
          file=sys.stderr)
    parquet.unlink(missing_ok=True)
    proc = subprocess.run([sys.executable, __file__, "_parquet", str(gz), str(parquet)])
    if proc.returncode == 0 and parquet.exists() and parquet.stat().st_size > 0:
        return True
    parquet.unlink(missing_ok=True)
    print(f"  parquet build failed (rc={proc.returncode}) — falling back to streaming mode.",
          file=sys.stderr)
    return False


def build(release: str, keep_gz: bool):
    cdir = cache_dir()
    gz = cdir / f"{release}-domain-ranks.txt.gz"
    parquet = cdir / f"{release}-domain-ranks.parquet"

    download_resumable(ranks_url(release), gz)

    # Count rows / find total (= max harmonicc_pos) with a single stdlib pass.
    print("Indexing (counting domains)...", file=sys.stderr)
    total = 0
    t0 = time.time()
    for row in iter_rows(gz):
        try:
            p = int(row[0])
            if p > total:
                total = p
        except ValueError:
            pass
    print(f"  {total:,} domains  ({time.time()-t0:.0f}s)", file=sys.stderr)

    # Write a WORKING streaming-mode index FIRST, so a later parquet crash can never
    # strand the build (the old failure mode on Py3.14 + duckdb).
    meta = {
        "release": release,
        "total_count": int(total),
        "gz": str(gz),
        "parquet": None,
        "ranks_url": ranks_url(release),
        "built_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
    }
    meta_path().write_text(json.dumps(meta, indent=2))
    print(f"Index ready (streaming .gz mode): {total:,} domains from {release}.", file=sys.stderr)

    # Best-effort upgrade to Parquet for sub-second repeat lookups. Crash-isolated.
    has_parquet = try_build_parquet(gz, parquet)
    if has_parquet:
        meta["parquet"] = str(parquet)
        meta_path().write_text(json.dumps(meta, indent=2))
        # Keep the .gz unless parquet exists AND user didn't ask to keep it.
        if not keep_gz:
            gz.unlink(missing_ok=True)
            print(f"Removed {gz.name} (Parquet built; pass --keep-gz to retain the .gz).",
                  file=sys.stderr)
        print(f"Upgraded index to parquet mode: {total:,} domains from {release}.",
              file=sys.stderr)


# ------------------------------------------------------------------------ lookup
def _row_to_record(domain, key, hpos, hval, ppos, pval, nh, total):
    return {
        "domain": domain, "host_rev": key,
        "hc_rank": int(hpos), "hc_value": float(hval),
        "hc_percentile": round(100 * (1 - int(hpos) / total), 4) if total else None,
        "hc_grade": grade(int(hpos)),
        "pr_rank": int(ppos), "pr_value": float(pval), "n_hosts": int(nh),
    }


def query_domains(domains):
    """Return (dict input_domain -> record|None, meta)."""
    meta = load_meta()
    total = meta["total_count"]
    rev = {d: reverse_domain(d) for d in domains}
    keys = {k for k in rev.values() if k}
    by_key = {}

    if meta.get("parquet") and Path(meta["parquet"]).exists() and _duckdb_usable():
        import duckdb
        con = duckdb.connect()
        ph = ",".join("?" for _ in keys)
        rows = con.execute(
            f"SELECT host_rev,harmonicc_pos,harmonicc_val,pr_pos,pr_val,n_hosts "
            f"FROM read_parquet(?) WHERE host_rev IN ({ph})",
            [meta["parquet"], *keys],
        ).fetchall()
        con.close()
        for hr, hpos, hval, ppos, pval, nh in rows:
            by_key[hr] = (hpos, hval, ppos, pval, nh)
    else:
        gz = Path(meta.get("gz", ""))
        if not gz.exists():
            sys.exit(f"Neither Parquet nor cached .gz found. Re-run: python hc_rank.py build")
        want = set(keys)
        for row in iter_rows(gz):
            hr = row[4]
            if hr in want:
                by_key[hr] = (row[0], row[1], row[2], row[3], row[5])
                want.discard(hr)
                if not want:
                    break

    out = {}
    for d, k in rev.items():
        r = by_key.get(k)
        out[d] = _row_to_record(d, k, *r, total) if r else None
    return out, meta


def cmd_lookup(domains):
    results, meta = query_domains(domains)
    print(f"# release={meta['release']}  domains_indexed={meta['total_count']:,}\n")
    hdr = f"{'DOMAIN':28} {'HC RANK':>13} {'GRADE':14} {'PCTILE':>8} {'PR RANK':>13} {'HOST_REV':24}"
    print(hdr)
    print("-" * len(hdr))
    for d in domains:
        r = results[d]
        if r is None:
            print(f"{d:28} {'NOT FOUND':>13} {'-':14} {'-':>8} {'-':>13} {reverse_domain(d):24}")
        else:
            pct = f"{r['hc_percentile']:.3f}%" if r['hc_percentile'] is not None else "-"
            print(f"{d:28} {r['hc_rank']:>13,} {r['hc_grade']:14} {pct:>8} "
                  f"{r['pr_rank']:>13,} {r['host_rev']:24}")
    print()


# -------------------------------------------------------------------------- bulk
def read_domain_list(arg_file, arg_domains):
    domains = []
    if arg_domains:
        domains += [d.strip() for d in arg_domains.split(",") if d.strip()]
    if arg_file:
        for line in Path(arg_file).read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#"):
                domains.append(line.split(",")[0].strip())
    seen, out = set(), []
    for d in domains:
        if d.lower() not in seen:
            seen.add(d.lower())
            out.append(d)
    return out


def cmd_bulk(arg_file, arg_domains, csv_out, xlsx_out):
    domains = read_domain_list(arg_file, arg_domains)
    if not domains:
        sys.exit("No domains given. Provide a file or --domains a,b,c")
    results, meta = query_domains(domains)
    cols = ["domain", "host_rev", "hc_rank", "hc_value", "hc_percentile",
            "hc_grade", "pr_rank", "pr_value", "n_hosts"]
    rows = []
    for d in domains:
        r = results[d]
        rows.append(r if r else {"domain": d, "host_rev": reverse_domain(d),
                                 "hc_rank": None, "hc_value": None, "hc_percentile": None,
                                 "hc_grade": "NOT FOUND", "pr_rank": None,
                                 "pr_value": None, "n_hosts": None})

    if xlsx_out:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "HC Rank"
        ws.append([c.upper() for c in cols])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for row in rows:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, w in enumerate([30, 26, 12, 16, 12, 16, 12, 16, 9], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(xlsx_out)
        print(f"Wrote {len(rows)} rows -> {xlsx_out}", file=sys.stderr)

    if csv_out or not xlsx_out:
        out = open(csv_out, "w", newline="", encoding="utf-8") if csv_out else sys.stdout
        w = csv.DictWriter(out, fieldnames=cols)
        w.writeheader()
        for row in rows:
            w.writerow({c: row.get(c) for c in cols})
        if csv_out:
            out.close()
            print(f"Wrote {len(rows)} rows -> {csv_out}", file=sys.stderr)

    found = sum(1 for d in domains if results[d])
    print(f"# {found}/{len(domains)} domains found in {meta['release']} "
          f"({meta['total_count']:,} indexed)", file=sys.stderr)


# ---------------------------------------------------------------------- releases
def cmd_releases():
    import urllib.request, re
    try:
        html = urllib.request.urlopen(WEBGRAPH_INDEX, timeout=60).read().decode("utf-8", "replace")
        rels = sorted(set(re.findall(r"cc-main-\d{4}-[a-z\-]+", html)), reverse=True)
    except Exception as e:
        rels = []
        print(f"(could not fetch listing: {e})", file=sys.stderr)
    print(f"Default / latest known: {LATEST_RELEASE}")
    if rels:
        print("Releases found in CC web-graphs index:")
        for r in rels[:20]:
            print(f"  {r}")
    print(f"\nListing page: {WEBGRAPH_INDEX}")
    print("Build a specific release:  python hc_rank.py build --release <id>")


# ---------------------------------------------------------------------- argparse
def main():
    p = argparse.ArgumentParser(
        description="Common Crawl Harmonic Centrality (HC Rank) lookup — bulk capable.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    sub = p.add_subparsers(dest="command", required=True)

    b = sub.add_parser("build", help="Download + index a web-graph release (one-time)")
    b.add_argument("--release", default=LATEST_RELEASE, help=f"default: {LATEST_RELEASE}")
    b.add_argument("--keep-gz", action="store_true", help="retain the raw .gz after indexing")

    lk = sub.add_parser("lookup", help="Look up HC Rank for one or more domains")
    lk.add_argument("domains", nargs="+")

    bk = sub.add_parser("bulk", help="Batch lookup -> CSV/XLSX")
    bk.add_argument("file", nargs="?", help="text file, one domain per line (# comments ok)")
    bk.add_argument("--domains", help="comma-separated domains (alt to file)")
    bk.add_argument("--csv", dest="csv_out")
    bk.add_argument("--xlsx", dest="xlsx_out")

    sub.add_parser("releases", help="List available web-graph releases")

    # Hidden: internal subprocess entry point for the crash-isolated parquet build.
    pq = sub.add_parser("_parquet", add_help=False)
    pq.add_argument("gz")
    pq.add_argument("parquet")

    args = p.parse_args()
    if args.command == "_parquet":
        _parquet_copy(Path(args.gz), Path(args.parquet))
    elif args.command == "build":
        build(args.release, args.keep_gz)
    elif args.command == "lookup":
        cmd_lookup(args.domains)
    elif args.command == "bulk":
        cmd_bulk(args.file, args.domains, args.csv_out, args.xlsx_out)
    elif args.command == "releases":
        cmd_releases()


if __name__ == "__main__":
    main()
